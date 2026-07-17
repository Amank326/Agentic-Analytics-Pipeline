import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
import sqlite3
import os

# Ensure directories exist
os.makedirs('public', exist_ok=True)

# =====================================================================
# 1. GENERATE SYNTHETIC DATASET (50+ rows with outliers)
# =====================================================================
np.random.seed(42)
n_rows = 65

sales = np.random.normal(loc=250, scale=90, size=n_rows)
profit = sales * np.random.uniform(0.12, 0.28, size=n_rows)

# Inject outliers manually (high sales, low profit/high profit)
sales[8] = 1400.00
profit[8] = -420.00

sales[22] = 1150.00
profit[22] = 450.00

sales[35] = 60.00
profit[35] = -180.00

# Base dates
dates = pd.date_range(start='2026-01-01', periods=n_rows).strftime('%Y-%m-%d').tolist()

categories = ['Technology', 'Furniture', 'Office Supplies']
sub_categories = {
    'Technology': ['Phones', 'Accessories', 'Copiers'],
    'Furniture': ['Chairs', 'Tables', 'Bookcases'],
    'Office Supplies': ['Art', 'Binders', 'Paper']
}
regions = ['East', 'West', 'Central', 'South']
customer_names = [f'Customer_{i}' for i in range(1, 21)]

data = []
for i in range(n_rows):
    cat = np.random.choice(categories)
    sub_cat = np.random.choice(sub_categories[cat])
    reg = np.random.choice(regions)
    cust = np.random.choice(customer_names)
    
    # Inject dirty data formatting for sales (e.g. string with $) in some rows
    sales_val = f"${sales[i]:.2f}" if i % 8 == 0 else sales[i]
    
    data.append({
        'order_id': i + 2000,
        'order_date': dates[i],
        'customer_name': cust,
        'category': cat,
        'sub_category': sub_cat,
        'sales': sales_val,
        'profit': profit[i],
        'quantity': np.random.randint(1, 10),
        'region': reg
    })

# Add explicit duplicates to test cleaning
data.append(data[4])
data.append(data[18])

df_raw = pd.DataFrame(data)

# =====================================================================
# 2. CLEANING PIPELINE
# =====================================================================
df_cleaned = df_raw.copy()

# Drop Duplicates
df_cleaned = df_cleaned.drop_duplicates()

# Clean sales strings
if df_cleaned['sales'].dtype == object:
    df_cleaned['sales'] = df_cleaned['sales'].astype(str).str.replace('$', '', regex=False)
df_cleaned['sales'] = pd.to_numeric(df_cleaned['sales'])

# Convert datatypes
df_cleaned['order_id'] = df_cleaned['order_id'].astype(int)
df_cleaned['order_date'] = pd.to_datetime(df_cleaned['order_date'])
df_cleaned['profit'] = pd.to_numeric(df_cleaned['profit'])
df_cleaned['quantity'] = df_cleaned['quantity'].astype(int)

# Sort by order ID
df_cleaned = df_cleaned.sort_values('order_id').reset_index(drop=True)

# Export clean datasets
df_cleaned.to_csv('cleaned_store_sales.csv', index=False)
df_cleaned.to_json('public/sales_data.json', orient='records', date_format='iso')

# =====================================================================
# 3. EXCEL WORKBOOK GENERATION (With openpyxl formatting & formulas)
# =====================================================================
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()

# Sheet 1: Store Sales Data
ws_data = wb.active
ws_data.title = "Store Sales Data"

# Write headers
headers = list(df_cleaned.columns)
ws_data.append(headers)

# Write data rows
for r in df_cleaned.itertuples(index=False):
    # Convert Timestamp to date string for excel compatibility
    row_val = list(r)
    row_val[1] = row_val[1].strftime('%Y-%m-%d')
    ws_data.append(row_val)

# Format Sheet 1
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)

for col_num in range(1, len(headers) + 1):
    cell = ws_data.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Format data rows
for row in range(2, len(df_cleaned) + 2):
    for col in range(1, len(headers) + 1):
        cell = ws_data.cell(row=row, column=col)
        cell.border = thin_border
        
        # Format Currency columns
        if col in [6, 7]:  # Sales and Profit columns
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal="right")
        elif col in [1, 8]:  # Order ID, Quantity
            cell.alignment = Alignment(horizontal="center")
        elif col == 2:  # Order Date
            cell.alignment = Alignment(horizontal="center")

# Auto-fit columns
for col in ws_data.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_data.column_dimensions[col_letter].width = max(max_len + 3, 11)

# Sheet 2: KPI Analytics Summary
ws_summary = wb.create_sheet(title="KPI Summary")

ws_summary.append(["Key Metrics Summary"])
ws_summary.append([])

ws_summary.append(["Metric", "Value"])
# Add formulas using cell ranges
last_row = len(df_cleaned) + 1
ws_summary.append(["Total Sales", f"=SUM('Store Sales Data'!F2:F{last_row})"])
ws_summary.append(["Total Profit", f"=SUM('Store Sales Data'!G2:G{last_row})"])
ws_summary.append(["Overall Profit Margin", "=B4/B3"])
ws_summary.append(["Average Order Value", f"=AVERAGE('Store Sales Data'!F2:F{last_row})"])
ws_summary.append(["Total Units Sold", f"=SUM('Store Sales Data'!H2:H{last_row})"])

# Format Summary Page
ws_summary.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
for col_num in range(1, 3):
    cell = ws_summary.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Styling metrics rows
for r in range(4, 9):
    ws_summary.cell(row=r, column=1).font = Font(bold=True)
    ws_summary.cell(row=r, column=1).border = thin_border
    cell_val = ws_summary.cell(row=r, column=2)
    cell_val.border = thin_border
    if r in [4, 5, 7]:
        cell_val.number_format = '$#,##0.00'
    elif r == 6:
        cell_val.number_format = '0.00%'
    elif r == 8:
        cell_val.number_format = '#,##0'

# Auto-fit column widths
ws_summary.column_dimensions['A'].width = 24
ws_summary.column_dimensions['B'].width = 16

# Save Excel files
wb.save('cleaned_store_sales.xlsx')
wb.save('public/cleaned_store_sales.xlsx')

# =====================================================================
# 4. SQLite DATABASE LOADER
# =====================================================================
conn = sqlite3.connect('SalesDB.db')
cursor = conn.cursor()

# Create table
cursor.execute('''
CREATE TABLE IF NOT EXISTS store_sales (
    order_id INTEGER PRIMARY KEY,
    order_date TEXT NOT NULL,
    customer_name TEXT NOT NULL,
    category TEXT NOT NULL,
    sub_category TEXT NOT NULL,
    sales REAL NOT NULL,
    profit REAL NOT NULL,
    quantity INTEGER NOT NULL,
    region TEXT NOT NULL
)
''')

# Replace and load cleaned data
df_sqlite = df_cleaned.copy()
df_sqlite['order_date'] = df_sqlite['order_date'].dt.strftime('%Y-%m-%d')
df_sqlite.to_sql('store_sales', conn, if_exists='replace', index=False)

# =====================================================================
# 5. EXECUTE BUSINESS ANALYTICS QUERIES
# =====================================================================
queries = {
    "Query 1: Category & Sub-Category Margin Report": '''
        SELECT 
            category,
            sub_category,
            SUM(sales) AS total_sales,
            SUM(profit) AS total_profit,
            (SUM(profit) / SUM(sales)) * 100 AS profit_margin_percent
        FROM store_sales
        GROUP BY category, sub_category
        ORDER BY total_sales DESC;
    ''',
    "Query 2: Monthly Performance Growth Trends": '''
        SELECT 
            strftime('%Y-%m', order_date) AS sales_month,
            SUM(sales) AS total_sales,
            SUM(profit) AS total_profit,
            COUNT(order_id) AS total_orders
        FROM store_sales
        GROUP BY sales_month
        ORDER BY sales_month ASC;
    ''',
    "Query 3: Customer Performance Ranking (Advanced Window)": '''
        SELECT 
            customer_name,
            SUM(sales) AS total_sales,
            SUM(profit) AS total_profit,
            DENSE_RANK() OVER (ORDER BY SUM(profit) DESC) as profitability_rank
        FROM store_sales
        GROUP BY customer_name
        LIMIT 10;
    '''
}

with open('query_results.txt', 'w') as f:
    for title, sql in queries.items():
        f.write(f"=== {title} ===\n")
        res_df = pd.read_sql_query(sql, conn)
        f.write(res_df.to_string(index=False))
        f.write("\n\n")

conn.close()

# =====================================================================
# 6. EXPLORATORY DATA ANALYSIS (EDA) VISUALIZATIONS
# =====================================================================
sns.set_theme(style="whitegrid")

# Chart 1: Sales vs Profit Outliers (Scatter)
plt.figure(figsize=(9, 5))
sns.scatterplot(
    data=df_cleaned, 
    x='sales', 
    y='profit', 
    hue='category', 
    size='quantity',
    sizes=(40, 400),
    alpha=0.8,
    palette='muted'
)
plt.axvline(x=800, color='red', linestyle='--', alpha=0.5, label='High Sales Threshold')
plt.axhline(y=0, color='black', linestyle='-', alpha=0.3)
plt.title('Sales vs Profit Outliers Analysis', fontsize=12, fontweight='bold', pad=12)
plt.xlabel('Sales ($)', fontsize=10)
plt.ylabel('Profit ($)', fontsize=10)
plt.legend(bbox_to_anchor=(1.02, 1), loc='upper left')
plt.tight_layout()
plt.savefig('sales_vs_profit_outliers.png', dpi=150)
plt.savefig('public/sales_vs_profit_outliers.png', dpi=150)
plt.close()

# Chart 2: Regional Sales performance by Category (Bar Chart)
plt.figure(figsize=(9, 5))
sns.barplot(
    data=df_cleaned,
    x='region',
    y='sales',
    hue='category',
    errorbar=None,
    palette='viridis'
)
plt.title('Average Sales Performance by Region & Product Category', fontsize=12, fontweight='bold', pad=12)
plt.xlabel('Region', fontsize=10)
plt.ylabel('Average Sales ($)', fontsize=10)
plt.legend(loc='upper right')
plt.tight_layout()
plt.savefig('sales_by_category_region.png', dpi=150)
plt.savefig('public/sales_by_category_region.png', dpi=150)
plt.close()

# Chart 3: Monthly Financial Trends (Line Chart)
plt.figure(figsize=(9, 5))
df_trend = df_cleaned.copy()
df_trend['month'] = df_trend['order_date'].dt.to_period('M').astype(str)
df_trend_grouped = df_trend.groupby('month')[['sales', 'profit']].sum().reset_index()

plt.plot(df_trend_grouped['month'], df_trend_grouped['sales'], marker='o', color='#4f46e5', linewidth=2.5, label='Monthly Sales')
plt.plot(df_trend_grouped['month'], df_trend_grouped['profit'], marker='s', color='#10b981', linewidth=2.5, label='Monthly Profit')
plt.title('Monthly Sales & Profit Growth Trajectory', fontsize=12, fontweight='bold', pad=12)
plt.xlabel('Sales Month', fontsize=10)
plt.ylabel('Amount ($)', fontsize=10)
plt.legend(loc='upper left')
plt.grid(True, linestyle='--', alpha=0.5)
plt.tight_layout()
plt.savefig('monthly_financial_trends.png', dpi=150)
plt.savefig('public/monthly_financial_trends.png', dpi=150)
plt.close()

# =====================================================================
# 7. EXECUTION COMPLETE LOGS
# =====================================================================
print("==========================================")
print("  Institutional Upgraded Pipeline Complete")
print("==========================================")
print(f"Raw rows parsed: {len(df_raw)}")
print(f"Duplicate rows removed: {len(df_raw) - len(df_cleaned)}")
print(f"Cleaned records loaded: {len(df_cleaned)}")
print("CSV dataset exported: cleaned_store_sales.csv")
print("Formatted Excel exported: cleaned_store_sales.xlsx")
print("SQLite database loaded: SalesDB.db")
print("Advanced SQL query logs written: query_results.txt")
print("EDA Chart 1 saved: sales_vs_profit_outliers.png")
print("EDA Chart 2 saved: sales_by_category_region.png")
print("EDA Chart 3 saved: monthly_financial_trends.png")
print("==========================================")
